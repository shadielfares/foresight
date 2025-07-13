import openpyxl

"""
Author: Shadi El-Fares
Purpose: Helper functions to extract data from the mentioned .xlsx file
Dated: 2025-05-27
"""

FILENAME = "well_known_people.xlsx"
WORKSPACE = openpyxl.load_workbook(filename=FILENAME)
SHEET = WORKSPACE["well_known_people.csv"]

def extract_name_col(sheet, arr):
    for row in sheet.iter_rows(min_col=1, min_row=2, max_row=sheet.max_row, values_only=True):
        arr.append(row[0])
    return arr

def extract_articles_col(sheet, arr):
    for row in sheet.iter_rows(min_col=3, min_row=2, max_row=sheet.max_row, values_only=True):
        arr.append(row[0])
    return arr